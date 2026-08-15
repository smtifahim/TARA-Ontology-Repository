"""
download_curated_sheets.py — Versioned Snapshot Downloader for TARA Curation Sheets

Downloads a specific historical revision of a curation Google Sheet, saves it
as a single .xlsx workbook with all of its original tabs, and augments the
workbook with provenance/navigation info:

  - A new first tab, "Google Sheet Source", recording the sheet's URI and the
    version that was downloaded.
  - A trailing "Source-Sheet-Row-Link" column on every original tab, with a
    clickable link back to the corresponding row on the live Google Sheet.
  - Frozen header row + first column on every tab.

Caveat: the Drive API does not expose the human-readable text typed into
"Name current version" in the Sheets UI, and `keepForever` (which the docs
describe for uploaded "blob" files) does not apply to native Sheets/Docs
revisions either — empirically, named versions come back with
keepForever=False just like every other revision. So every revision is
listed and selected by timestamp (+ last-modified-by), newest first; use
Sheets > File > Version history > See version history to match a named
version's timestamp to the right entry in the prompt.

Caveat 2: Google's own Sheets -> XLSX export has a bug where a cell whose
text is only partially hyperlinked (e.g. a citation ending in a plain-text
URL that Sheets auto-linked) gets exported with that hyperlinked run's
literal text dropped, leaving only a whole-cell hyperlink behind. This
script detects and repairs those specific cells after download by
cross-checking against the live sheet's true text via the Sheets API
(get_live_hyperlinked_cell_texts).

Setup
-----
1. pip install -r requirements.txt
2. Create an OAuth "Desktop app" client in Google Cloud Console, download its
   client secret JSON, and save it as curated-data/.credentials/client_secret.json
   (the account must have at least read access to the target sheets).
3. Run this script; the first run opens a browser for consent and caches a
   token at curated-data/.credentials/token.json for subsequent runs.

Usage
-----
  python download_curated_sheets.py

Author: Fahim Imam
"""

import os
import re
from datetime import datetime

from google.auth.transport.requests import AuthorizedSession, Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Directories / paths
# ---------------------------------------------------------------------------
_HERE = os.path.dirname(os.path.abspath(__file__))
CREDENTIALS_DIR = os.path.join(_HERE, '.credentials')
CLIENT_SECRET_FILE = os.path.join(CREDENTIALS_DIR, 'client_secret.json')
TOKEN_FILE = os.path.join(CREDENTIALS_DIR, 'token.json')

SCOPES = [
    'https://www.googleapis.com/auth/drive.readonly',
    'https://www.googleapis.com/auth/spreadsheets.readonly',
]

XLSX_MIME_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

# ---------------------------------------------------------------------------
# Configurable data dictionary — source Google Sheets and where each one's
# downloaded workbook should be saved (paths are relative to this directory).
# ---------------------------------------------------------------------------
SHEETS_CONFIG = {
    'ontology-curation-sheet': {
        'label': 'TARA Ontology Curation Sheet',
        'sheet_id': '1hvUcTrw-b9ly8Yn1P706px22li0vsjslukYhxkTDlA8',
        'output_dir': 'ontology-curation-sheet',
    },
    'kb-curation-sheet-main': {
        'label': 'TARA KB Curation Sheet',
        'sheet_id': '14WIm3G6PoIHYrECjEetgd7pvKjzrntR57HbdMJ2ZNQI',
        'output_dir': 'kb-curation-sheets/sources',
    },
    # Additional KB source sheets can be added here as new entries, e.g.:
    # 'kb-curation-sheet-2': {
    #     'label': 'Some Additional KB Source Sheet',
    #     'sheet_id': '<google-sheet-id>',
    #     'output_dir': 'kb-curation-sheets/sources',
    # },
}


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
def get_credentials() -> Credentials:
    """Load cached OAuth credentials, refreshing or re-authorizing as needed."""
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(CLIENT_SECRET_FILE):
                raise FileNotFoundError(
                    f"Missing OAuth client secret file: {CLIENT_SECRET_FILE}\n"
                    "Create an OAuth Desktop-app client in Google Cloud Console "
                    "and save its client secret JSON there."
                )
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
            creds = flow.run_local_server(port=0)

        os.makedirs(CREDENTIALS_DIR, exist_ok=True)
        with open(TOKEN_FILE, 'w') as f:
            f.write(creds.to_json())

    return creds


# ---------------------------------------------------------------------------
# Revision discovery + selection
#
# Note: `keepForever` only applies to uploaded "blob" files (images, PDFs,
# etc.) per Google's own Drive API docs — it is not how native Google Sheets
# "Name current version" is exposed via the API. There is no field in the
# public Drive API that surfaces the name text you typed in the Sheets UI, so
# every revision is listed here (newest first) and you cross-reference the
# timestamp against Sheets > File > Version history > See version history,
# which does show your named labels next to their timestamps.
# ---------------------------------------------------------------------------
def list_all_revisions(drive_service, sheet_id: str) -> list:
    """Return every revision for a sheet (paginated), newest first."""
    all_revisions = []
    page_token = None
    while True:
        resp = drive_service.revisions().list(
            fileId=sheet_id,
            fields='nextPageToken, revisions(id,modifiedTime,keepForever,lastModifyingUser,exportLinks)',
            pageToken=page_token,
        ).execute()
        all_revisions.extend(resp.get('revisions', []))
        page_token = resp.get('nextPageToken')
        if not page_token:
            break
    all_revisions.sort(key=lambda r: r['modifiedTime'], reverse=True)
    return all_revisions


def _format_revision_label(revision: dict) -> str:
    modified = revision['modifiedTime']
    try:
        dt = datetime.strptime(modified, '%Y-%m-%dT%H:%M:%S.%fZ')
    except ValueError:
        dt = datetime.strptime(modified, '%Y-%m-%dT%H:%M:%SZ')
    timestamp = dt.strftime('%Y-%m-%d %H:%M UTC')
    editor = revision.get('lastModifyingUser', {}).get('displayName', 'unknown user')
    return f"{timestamp} — edited by {editor} (revision {revision['id']})"


def prompt_for_revision(revisions: list, sheet_label: str) -> dict:
    """Prompt the user to pick one revision from a list, newest first.

    Cross-reference the printed timestamps against Sheets > File > Version
    history > See version history to identify which entry is the named
    version you want (the Drive API can't tell us the name text directly).
    """
    if not revisions:
        raise RuntimeError(f"No revisions found for '{sheet_label}'.")

    print(f"\nRevisions available for '{sheet_label}' (newest first):")
    for i, revision in enumerate(revisions, start=1):
        print(f"  [{i}] {_format_revision_label(revision)}")
    print(
        "Check Sheets > File > Version history > See version history to match "
        "a named version's timestamp to one of the entries above."
    )

    while True:
        choice = input(f"Select a version to download [1-{len(revisions)}]: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(revisions):
            return revisions[int(choice) - 1]
        print("Invalid selection, try again.")


# ---------------------------------------------------------------------------
# Live sheet metadata (tab name -> gid, sheet title)
# ---------------------------------------------------------------------------
def get_tab_gid_map(sheets_service, sheet_id: str) -> dict:
    """Return {tab_title: gid} for every tab in the live sheet."""
    meta = sheets_service.spreadsheets().get(
        spreadsheetId=sheet_id, fields='sheets.properties(title,sheetId)'
    ).execute()
    return {
        s['properties']['title']: s['properties']['sheetId']
        for s in meta.get('sheets', [])
    }


def get_sheet_name(drive_service, sheet_id: str) -> str:
    """Return the live Google Sheet's exact file name (title)."""
    meta = drive_service.files().get(fileId=sheet_id, fields='name').execute()
    return meta['name']


def get_live_hyperlinked_cell_texts(sheets_service, sheet_id: str) -> dict:
    """Return {tab_title: {(row, col): formattedValue}} for every live cell
    that carries a hyperlink.

    Works around a Google Sheets -> XLSX export bug: when only part of a
    cell's text is hyperlinked (e.g. a citation ending in a plain-text URL
    that Sheets auto-linked), the exported .xlsx drops that hyperlinked run's
    literal text entirely and reattaches the hyperlink to whatever text is
    left. The Sheets API's `formattedValue`, unlike the xlsx export, always
    reflects the true underlying text, so it's used here to repair just the
    affected cells post-download.
    """
    resp = sheets_service.spreadsheets().get(
        spreadsheetId=sheet_id,
        includeGridData=True,
        fields='sheets(properties(title),data(rowData(values(formattedValue,hyperlink))))',
    ).execute()

    result = {}
    for sheet in resp.get('sheets', []):
        title = sheet['properties']['title']
        cell_map = {}
        data = sheet.get('data') or [{}]
        for row_idx, row in enumerate(data[0].get('rowData', []), start=1):
            for col_idx, cell in enumerate(row.get('values', []), start=1):
                if cell.get('hyperlink'):
                    cell_map[(row_idx, col_idx)] = cell.get('formattedValue', '')
        result[title] = cell_map
    return result


# ---------------------------------------------------------------------------
# Download a specific revision as .xlsx
# ---------------------------------------------------------------------------
def download_revision_xlsx(credentials: Credentials, revision: dict, dest_path: str) -> None:
    export_link = revision.get('exportLinks', {}).get(XLSX_MIME_TYPE)
    if not export_link:
        raise RuntimeError(
            f"Revision {revision['id']} has no xlsx export link available."
        )
    session = AuthorizedSession(credentials)
    resp = session.get(export_link)
    resp.raise_for_status()
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    with open(dest_path, 'wb') as f:
        f.write(resp.content)


# ---------------------------------------------------------------------------
# Post-process the downloaded workbook
# ---------------------------------------------------------------------------
_FORMULA_PREFIX_RE = re.compile(r'^[=+\-@|]')


def _row_link_url(sheet_id: str, gid, row_number: int) -> str:
    # A full-row range (e.g. "5:5") selects and scrolls to the entire row,
    # instead of just the first cell.
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit#gid={gid}&range={row_number}:{row_number}"


def postprocess_workbook(
    path: str,
    sheet_id: str,
    sheet_uri: str,
    version_label: str,
    tab_gid_map: dict,
    live_hyperlinked_cell_texts: dict = None,
) -> None:
    wb = load_workbook(path)
    original_sheet_names = list(wb.sheetnames)
    live_hyperlinked_cell_texts = live_hyperlinked_cell_texts or {}

    # --- Repair cells whose hyperlinked-run text was dropped by Google's own
    # xlsx export (see get_live_hyperlinked_cell_texts) ---
    repaired_count = 0
    for sheet_name in original_sheet_names:
        ws = wb[sheet_name]
        cell_map = live_hyperlinked_cell_texts.get(sheet_name, {})
        for (row_number, col_number), correct_text in cell_map.items():
            if row_number > ws.max_row or col_number > ws.max_column:
                continue
            cell = ws.cell(row=row_number, column=col_number)
            if cell.value != correct_text:
                cell.value = correct_text
                repaired_count += 1
    if repaired_count:
        print(
            f"  Repaired {repaired_count} cell(s) whose hyperlinked text was "
            "dropped by Google's xlsx export (restored from the live sheet)."
        )

    # --- Source-Sheet-Row-Link column + frozen panes on every original tab ---
    for sheet_name in original_sheet_names:
        ws = wb[sheet_name]
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        gid = tab_gid_map.get(sheet_name)
        link_col = ws.max_column + 1
        ws.cell(row=1, column=link_col, value='Source-Sheet-Row-Link')

        if gid is not None:
            for row_number in range(2, ws.max_row + 1):
                url = _row_link_url(sheet_id, gid, row_number)
                cell = ws.cell(row=row_number, column=link_col, value=url)
                cell.hyperlink = url
                cell.style = 'Hyperlink'

        ws.freeze_panes = 'B2'

    # --- New first tab: Google Sheet Source ---
    source_ws = wb.create_sheet('Google Sheet Source', 0)
    source_ws.append(['Google Sheet URI', 'Named Version'])
    source_ws.append([sheet_uri, version_label])
    source_ws.freeze_panes = 'B2'

    wb.save(path)


# ---------------------------------------------------------------------------
# Core download routine shared by both public entry points
# ---------------------------------------------------------------------------
def _download_sheet_version(config: dict, credentials, drive_service, sheets_service) -> str:
    sheet_id = config['sheet_id']
    label = config['label']

    revisions = list_all_revisions(drive_service, sheet_id)
    revision = prompt_for_revision(revisions, label)
    version_label = _format_revision_label(revision)

    sheet_name = get_sheet_name(drive_service, sheet_id)
    tab_gid_map = get_tab_gid_map(sheets_service, sheet_id)
    live_hyperlinked_cell_texts = get_live_hyperlinked_cell_texts(sheets_service, sheet_id)
    sheet_uri = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit"

    output_dir = os.path.join(_HERE, config['output_dir'])
    dest_path = os.path.join(output_dir, f"{sheet_name}.xlsx")

    print(f"Downloading '{label}' — {version_label} ...")
    download_revision_xlsx(credentials, revision, dest_path)
    postprocess_workbook(
        dest_path, sheet_id, sheet_uri, version_label, tab_gid_map,
        live_hyperlinked_cell_texts,
    )
    # print(f"  Saved: {dest_path}")
    print(f"  Saved: {os.path.relpath(dest_path)}")

    return dest_path


# ---------------------------------------------------------------------------
# Public functions
# ---------------------------------------------------------------------------
def download_ontology_curation_sheet() -> str:
    """Download a named version of the TARA Ontology Curation Sheet as .xlsx."""
    credentials = get_credentials()
    drive_service = build('drive', 'v3', credentials=credentials)
    sheets_service = build('sheets', 'v4', credentials=credentials)
    config = SHEETS_CONFIG['ontology-curation-sheet']
    return _download_sheet_version(config, credentials, drive_service, sheets_service)


def download_kb_curation_sheet(sheet_keys=None) -> list:
    """Download a named version of one or more KB curation source sheets as .xlsx.

    sheet_keys: optional list of SHEETS_CONFIG keys to download. If omitted,
    the user is prompted to choose one, several, or all configured KB sheets.
    """
    kb_keys = [k for k in SHEETS_CONFIG if k.startswith('kb-')]

    if sheet_keys is None:
        print("\nConfigured KB curation source sheets:")
        for i, key in enumerate(kb_keys, start=1):
            print(f"  [{i}] {SHEETS_CONFIG[key]['label']}")
        choice = input(
            f"Select sheet(s) to download (comma-separated indices, or 'all') [1-{len(kb_keys)}]: "
        ).strip().lower()

        if choice == 'all':
            sheet_keys = kb_keys
        else:
            indices = [int(x) for x in re.split(r'[,\s]+', choice) if x.isdigit()]
            sheet_keys = [kb_keys[i - 1] for i in indices if 1 <= i <= len(kb_keys)]

        if not sheet_keys:
            print("No valid selection made; nothing downloaded.")
            return []

    credentials = get_credentials()
    drive_service = build('drive', 'v3', credentials=credentials)
    sheets_service = build('sheets', 'v4', credentials=credentials)

    dest_paths = []
    for key in sheet_keys:
        config = SHEETS_CONFIG[key]
        dest_paths.append(_download_sheet_version(config, credentials, drive_service, sheets_service))

    return dest_paths


# ---------------------------------------------------------------------------
# CLI menu
# ---------------------------------------------------------------------------
def main():
    print("=" * 70)
    print("TARA Curated Sheets Downloader")
    print("=" * 70)
    print("  [1] Ontology Curation Sheet")
    print("  [2] KB Curation Sheet(s)")
    choice = input("Select what to download [1-2]: ").strip()

    if choice == '1':
        download_ontology_curation_sheet()
    elif choice == '2':
        download_kb_curation_sheet()
    else:
        print("Invalid selection.")


if __name__ == '__main__':
    main()
